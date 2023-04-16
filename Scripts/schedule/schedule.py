import requests
import json

from schedule import constants as sc
from openpyxl import load_workbook
from os import path, remove, listdir
from bs4 import BeautifulSoup


def get_legacy_file_path():
    return path.join(sc.SCHEDULE_XLSX_PATH, sc.LEGACY_FILENAME)


def get_url_data_path():
    return path.join(sc.SCHEDULE_DATA_PATH, sc.URL_DATA_FILENAME)


def get_iit_data_path():
    return path.join(sc.SCHEDULE_DATA_PATH, sc.IIT_DATA_FILENAME)


def is_legacy_file():
    return path.isfile(get_legacy_file_path())


def is_url_data():
    return path.isfile(get_url_data_path())


def is_iit_data():
    return path.isfile(get_iit_data_path())


def get_available_weeks_data_path():
    return path.join(sc.SCHEDULE_DATA_PATH, sc.AVAILABLE_WEEKS_DATA_FILENAME)


def is_available_weeks_data():
    return path.isfile(get_available_weeks_data_path())


def clear_xlsx_files():
    for file in listdir(sc.SCHEDULE_XLSX_PATH):
        remove(path.join(sc.SCHEDULE_XLSX_PATH, file))


def clear_weekly_schedule_files():
    for file in listdir(sc.WEEKLY_SCHEDULE_DATA_PATH):
        remove(path.join(sc.WEEKLY_SCHEDULE_DATA_PATH, file))


class Schedule:
    def __init__(self):
        self.url = None
        self.iit_schedules_data = dict()
        self.available_weeks = list()

    def check_url(self):
        response = requests.get(sc.UNIVERS_SCHEDULE)
        if response.status_code == 200:
            schedules_html = BeautifulSoup(response.content, 'html.parser')
            semester_block = schedules_html.find(class_='document-accordion document-accordion-2',
                                                 id='bx_89755312_75167')

            file_names = semester_block.find_all('div', class_='document-link__name')
            file_urls = semester_block.find_all('a')

            filtered_names = [name.text.strip() for name in file_names]
            filtered_urls = [url.get('href') for url in file_urls]
            self.iit_schedules_data = dict(zip(filtered_names, filtered_urls))

            new_url = f'https://www.sevsu.ru{self.iit_schedules_data.get("ИИТ-21-о")}'
            if new_url != self.url:
                self.url = new_url
                return True
            return False

    def download(self):
        response = requests.get(self.url)  # запрос на файл по ссылке
        if response.status_code == 200:
            clear_xlsx_files()  # удаление существующих xlsx файлов расписаний
            clear_weekly_schedule_files()  # удаление существующих txt файлов расписаний
            with open(get_legacy_file_path(), 'wb') as file:
                file.write(response.content)
            return True
        return False

    def check_to_download(self):
        url_updated = self.check_url()
        return not is_legacy_file() or url_updated

    def define_available_weeks(self):
        self.available_weeks.clear()
        book = load_workbook(get_legacy_file_path(), read_only=True)

        for sheetname in book.sheetnames:
            week_name = str(sheetname).partition('(')[0].partition(' ')[2]
            if week_name.isdigit():
                self.available_weeks.append(int(week_name))

    def save_iit_schedules_data(self):
        # Сохранение данных расписания института в файл
        with open(get_iit_data_path(), 'w') as file:
            json.dump(self.iit_schedules_data, file)

    def save_url_data(self):
        # Сохранение данных о ссылке на файл расписания
        with open(get_url_data_path(), 'w') as file:
            file.write(self.url)

    def save_available_weeks_data(self):
        # Сохранение данных о доступных неделях расписания
        with open(get_available_weeks_data_path(), 'w') as file:
            for w in self.available_weeks:
                file.write(f'{w}\n')

    def load_url(self):
        # Загрузка данных о ссылке
        if is_url_data():
            with open(get_url_data_path(), 'r') as file:
                self.url = file.readline().rstrip()

    def load_available_weeks(self):
        self.available_weeks.clear()
        # Загрузка данных о доступных неделях
        if is_available_weeks_data():
            with open(get_available_weeks_data_path(), 'r') as file:
                for w in file:
                    self.available_weeks.append(int(w))


def get_last_message_data_path():
    return path.join(sc.SCHEDULE_DATA_PATH, sc.LAST_MESSAGE_DATA_FILE_NAME)


def is_last_message_data():
    return path.isfile(get_last_message_data_path())


class LastScheduleMessage:
    def __init__(self, message_id = None, chat_id = None):
        self.message_id = message_id
        self.chat_id = chat_id

    def save_message_data(self):
        with open(get_last_message_data_path(), 'w') as file:
            file.write(str(self.message_id) + '\n')
            file.write(str(self.chat_id))

    def load_message_data(self):
        if is_last_message_data():
            with open(get_last_message_data_path(), 'r') as file:
                self.message_id = int(file.readline().rstrip())
                self.chat_id = int(file.readline().rstrip())
