import schedule_constants
import schedule
import json

from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import path


def get_filename(week_num, group_num, default_type=True):
    filename = f'{schedule_constants.IIT_GROUP_NAMES.get(group_num)} ({week_num} нед.)'
    filename += '.json' if default_type else '.xlsx'
    return filename


def get_week():
    return datetime.today().isocalendar().week


def get_xlsx_file_path(filename):
    return path.join(schedule_constants.XLSX_PATH, filename)


def get_weekly_schedule_data_path(filename):
    return path.join(schedule_constants.WEEKLY_SCHEDULE_DATA_PATH, filename)


def get_available_weeks_data_path():
    return path.join(schedule_constants.DATA_PATH, schedule_constants.AVAILABLE_WEEKS_DATA_FILENAME)


def is_available_weeks_data():
    return path.isfile(get_available_weeks_data_path())


def is_xlsx_file(filename):
    return path.isfile(get_xlsx_file_path(filename))


def is_weekly_schedule_data(filename):
    return path.isfile(get_weekly_schedule_data_path(filename))


class WeeklySchedule:
    def __init__(self, group_num=2, week_num=get_week(), day_num=3):
        self.group_num = group_num
        self.week_num = week_num
        self.day_num = day_num
        self.weekly_schedule = dict()

    def get_weekly_schedule_day(self, day_num=3):
        return self.weekly_schedule.get(schedule_constants.DAY_NAMES.get(day_num))

    def check_to_create_new(self):
        filename = get_filename(week_num=self.week_num, group_num=self.group_num)
        return is_weekly_schedule_data(filename)

    def create_xlsx(self):
        book = load_workbook(schedule.get_legacy_file_path(), read_only=True)
        sheet_name = f'неделя {self.week_num}(уч.н.{self.week_num + 18})'
        sheet = book[sheet_name]

        # новый excel файл
        weekly_book = Workbook()
        weekly_sheet = weekly_book.active
        weekly_sheet = self.copy_data(sheet, weekly_sheet)  # копирование данных из по выбранной недели

        filename = get_filename(group_num=self.group_num, week_num=self.week_num, default_type=False)
        weekly_book.save(get_xlsx_file_path(filename))  # сохранение изменений

    def copy_data(self, sheet, weekly_sheet):
        group_name = schedule_constants.IIT_GROUP_NAMES.get(self.group_num)
        beg_column, column_len, beg_row, end_row, cur_column = 1, 10, 4 if self.week_num % 2 != 0 else 1, 54, 1

        while True:
            cell = sheet.cell(row=beg_row, column=cur_column)
            cell_text = str(cell.value).strip().replace('/', '')
            if group_name in cell_text:
                beg_column = cur_column
                break
            cur_column += 1

        # копирование данных по срезу
        for r in range(beg_row, end_row + 1):
            for c in range(beg_column, beg_column + column_len):
                legacy_value = sheet.cell(row=r, column=c).value
                weekly_sheet.cell(row=r - beg_row + 1, column=c - beg_column + 1, value=legacy_value)
        return weekly_sheet

    def define_weekly_schedule(self):
        self.weekly_schedule.clear()

        filename = get_filename(group_num=self.group_num, week_num=self.week_num, default_type=False)
        weekly_book = load_workbook(get_xlsx_file_path(filename=filename), read_only=True)
        improve_sheet = weekly_book.active

        day_num = 1
        for row in range(4, 52, 8):
            day_schedule_value = f'🔎 *{schedule_constants.IIT_GROUP_NAMES.get(self.group_num)} | {self.week_num} Неделя | {schedule_constants.DAY_NAMES.get(day_num)}*'

            for pair_num in range(row, row + 8):
                time_block = improve_sheet[f'D{pair_num}'].value

                pair1 = improve_sheet[f'E{pair_num}'].value
                pair2 = improve_sheet[f'H{pair_num}'].value

                pair_type1 = improve_sheet[f'F{pair_num}'].value
                pair_type2 = improve_sheet[f'I{pair_num}'].value

                pair_location1 = improve_sheet[f'G{pair_num}'].value
                pair_location2 = improve_sheet[f'J{pair_num}'].value

                if pair1 or pair2:
                    pair_text = f'\n🛌 *{pair_num - row + 1} Пара* `[{time_block}]`'

                    # если пара общая
                    if pair1 and not pair2 and (pair_location2 or pair_type2):
                        pair_text += f'\n{str(pair1).strip().replace("подгр.:0(из 2),", "")}'
                        fixed_type_pair = str(pair_type2).strip().replace('\n', ' ')
                        pair_text += f' ({fixed_type_pair})'
                    else:
                        if pair1:  # если пара для 1-ой подгруппы
                            pair_text += '*\n1 Подгруппа*'
                            pair_text += f'\n{str(pair1).strip().replace("подгр.:1(из 2),", "")}'
                            pair_text += f' ({str(pair_type1).strip()})' if pair_type1 else ''
                            pair_text += f' ({str(pair_location1).strip()})' if pair_location1 else ''
                        if pair2:  # если пара для 2-ой подгруппы
                            pair_text += '\n*2 Подгруппа*'
                            pair_text += f'\n{str(pair2).strip().replace("подгр.:2(из 2),", "")}'
                            pair_text += f' ({str(pair_type2).strip()})' if pair_type2 else ''
                            pair_text += f' ({str(pair_location2).strip()})' if pair_location2 else ''

                    day_schedule_value += f'\n{pair_text}'

            self.weekly_schedule[schedule_constants.DAY_NAMES.get(day_num)] = day_schedule_value
            day_num += 1

    def save_data(self):
        filename = get_filename(week_num=self.week_num, group_num=self.group_num)
        with open(get_weekly_schedule_data_path(filename), 'w') as file:
            json.dump(self.weekly_schedule, file)

    def load_data(self):
        filename = get_filename(week_num=self.week_num, group_num=self.group_num)
        with open(get_weekly_schedule_data_path(filename), 'r') as file:
            self.weekly_schedule.clear()
            self.weekly_schedule = json.load(file)
